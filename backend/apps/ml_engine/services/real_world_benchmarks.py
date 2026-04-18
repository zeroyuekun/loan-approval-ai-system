"""Real-world Australian lending benchmarks — distribution-level calibration data.

Fetches distribution-level data from public Australian sources to calibrate
the synthetic data generator against real-world lending patterns.

This complements MacroDataService (which provides point-in-time macro
indicators like cash rate, unemployment, property growth) with distribution
parameters (income percentiles, loan sizes, default rates, credit scores).

Data sources (all free, no commercial license required):
- ABS Data API (data.api.abs.gov.au): income distributions, lending indicators
- RBA Statistical Tables (rba.gov.au): lending rates, household debt
- APRA Quarterly ADI Statistics (apra.gov.au): arrears, LVR/DTI distributions
- Equifax published scorecards: credit score distributions by age/state

Every method has hardcoded fallback values matching the current DataGenerator
constants, so generation never breaks if APIs are unreachable.
"""

import csv
import io
import logging
from datetime import UTC, datetime, timedelta

import httpx

logger = logging.getLogger(__name__)

_CACHE_TTL_HOURS = 168  # 7 days — this data updates quarterly at most


# =====================================================================
# Fallback values — match current DataGenerator hardcoded constants
# exactly so that failed fetches produce identical generation output.
# =====================================================================

_FALLBACK_INCOME_PERCENTILES = {
    # ATO Taxation Statistics 2022-23 + ABS Employee Earnings Aug 2025
    "national": {
        "P10": 25_000,
        "P25": 42_000,
        "P50": 74_100,
        "P75": 105_000,
        "P90": 155_000,
        "source": "fallback (ATO 2022-23 / ABS Aug 2025)",
    },
    "NSW": {"P10": 27_000, "P25": 45_360, "P50": 80_028, "P75": 113_400, "P90": 167_400},
    "VIC": {"P10": 25_750, "P25": 43_260, "P50": 76_323, "P75": 108_150, "P90": 159_650},
    "QLD": {"P10": 23_750, "P25": 39_900, "P50": 70_395, "P75": 99_750, "P90": 147_250},
    "WA": {"P10": 28_000, "P25": 47_040, "P50": 82_992, "P75": 117_600, "P90": 173_600},
    "SA": {"P10": 23_000, "P25": 38_640, "P50": 68_172, "P75": 96_600, "P90": 142_600},
    "TAS": {"P10": 22_000, "P25": 36_960, "P50": 65_208, "P75": 92_400, "P90": 136_400},
    "ACT": {"P10": 31_250, "P25": 52_500, "P50": 92_625, "P75": 131_250, "P90": 193_750},
    "NT": {"P10": 26_250, "P25": 44_100, "P50": 77_805, "P75": 110_250, "P90": 162_750},
}

_FALLBACK_AVG_LOAN_SIZES = {
    # ABS Lending Indicators Dec Q 2025
    "owner_occupier": 693_801,
    "first_home_buyer": 560_249,
    "investor": 685_634,
    "personal": 24_500,
    "business": 45_000,
    "source": "fallback (ABS Lending Indicators Dec Q 2025)",
}

_FALLBACK_APRA_ARREARS = {
    # APRA Quarterly ADI Property Exposures Sep Q 2025
    "npl_rate": 0.0104,
    "total_arrears_rate": 0.0091,
    "lvr_80_plus_pct": 0.308,
    "dti_6_plus_pct": 0.061,
    "by_lvr_band": {
        "0_60": {"share": 0.35, "arrears_rate": 0.003},
        "60_70": {"share": 0.18, "arrears_rate": 0.005},
        "70_80": {"share": 0.16, "arrears_rate": 0.008},
        "80_90": {"share": 0.22, "arrears_rate": 0.012},
        "90_95": {"share": 0.09, "arrears_rate": 0.018},
    },
    "by_dti_band": {
        "0_4": {"share": 0.58, "arrears_rate": 0.004},
        "4_6": {"share": 0.35, "arrears_rate": 0.009},
        "6_plus": {"share": 0.061, "arrears_rate": 0.022},
    },
    "quarter": "Q3_2025",
    "source": "fallback (APRA Sep Q 2025)",
}

_FALLBACK_LENDING_RATES = {
    # RBA Table F5 — Indicator Lending Rates (approx 2025/2026)
    "avg_variable_rate": 6.50,
    "avg_fixed_1yr": 6.20,
    "avg_fixed_3yr": 5.90,
    "discounted_rate": 6.10,
    "source": "fallback (RBA F5 approx 2025/2026)",
}

# Equifax 2025 Credit Scorecard — published quarterly, no public API.
# Values transcribed from latest published release.
_FALLBACK_CREDIT_SCORE_DISTRIBUTIONS = {
    "18_30": {"mean": 715, "std": 95},
    "31_40": {"mean": 839, "std": 75},
    "41_50": {"mean": 872, "std": 65},
    "51_60": {"mean": 892, "std": 55},
    "61_plus": {"mean": 910, "std": 50},
    "national_mean": 864,
    "source": "fallback (Equifax 2025 Scorecard)",
}

_FALLBACK_APPROVAL_RATES = {
    "big_4": 0.65,
    "non_major": 0.60,
    "non_bank": 0.72,
    "overall": 0.65,
    "source": "fallback (APRA/ABS derived)",
}

_FALLBACK_SA4_UNEMPLOYMENT = {
    # ABS Labour Force Survey SA4-level unemployment rates (2025)
    # SA4 code → unemployment rate (proportion)
    "101": 0.034,  # Sydney - City and Inner South
    "102": 0.038,  # Sydney - Eastern Suburbs
    "104": 0.042,  # Sydney - Ryde
    "106": 0.042,  # Hunter Valley exc Newcastle
    "110": 0.045,  # Sydney - Outer South West
    "116": 0.048,  # Sydney - South West
    "117": 0.039,  # Sydney - Parramatta
    "201": 0.036,  # Melbourne - Inner
    "202": 0.040,  # Geelong
    "206": 0.044,  # Melbourne - North East
    "210": 0.046,  # Melbourne - Outer East
    "213": 0.052,  # Melbourne - West
    "301": 0.038,  # Brisbane Inner City
    "305": 0.041,  # Brisbane - South
    "309": 0.045,  # Gold Coast
    "313": 0.043,  # Sunshine Coast
    "315": 0.039,  # Sunshine Coast (hinterland)
    "401": 0.041,  # Adelaide - Central and Hills
    "501": 0.032,  # Perth - Inner
    "601": 0.040,  # Hobart
    "701": 0.035,  # Darwin
    "801": 0.028,  # Australian Capital Territory
    "source": "fallback (ABS Labour Force 2025)",
}

_FALLBACK_INDUSTRY_INCOME_MULTIPLIERS = {
    # ABS Average Weekly Earnings by ANZSIC division, relative to national median
    "A": 0.78,  # Agriculture, Forestry and Fishing
    "B": 1.85,  # Mining
    "C": 0.95,  # Manufacturing
    "D": 1.25,  # Electricity, Gas, Water and Waste Services
    "E": 1.10,  # Construction
    "F": 1.05,  # Wholesale Trade
    "G": 0.68,  # Retail Trade
    "H": 0.58,  # Accommodation and Food Services
    "I": 1.05,  # Transport, Postal and Warehousing
    "J": 1.40,  # Information Media and Telecommunications
    "K": 1.45,  # Financial and Insurance Services
    "L": 1.00,  # Rental, Hiring and Real Estate Services
    "M": 1.35,  # Professional, Scientific and Technical Services
    "N": 0.75,  # Administrative and Support Services
    "O": 1.20,  # Public Administration and Safety
    "P": 0.92,  # Education and Training
    "Q": 0.88,  # Health Care and Social Assistance
    "R": 0.72,  # Arts and Recreation Services
    "S": 0.80,  # Other Services
    "source": "fallback (ABS AWE by ANZSIC 2025)",
}

_FALLBACK_F6_RATES = {
    # RBA Table F6 — Housing Lending Rates (approx 2025/2026)
    "owner_occupier_variable": 6.27,
    "owner_occupier_fixed_3yr": 5.89,
    "investor_variable": 6.58,
    "investor_fixed_3yr": 6.15,
    "source": "fallback (RBA F6 approx 2025/2026)",
}

_FALLBACK_RBA_HOUSEHOLD_DEBT = {
    # RBA Table E2 — Household Finances: Selected Ratios (Dec Q 2025)
    "housing_debt_to_income": 1.41,
    "total_debt_to_income": 1.87,
    "debt_to_assets": 0.20,
    "interest_payments_to_income": 0.098,
    "quarter": "Q4_2025",
    "source": "fallback (RBA E2 Dec Q 2025)",
}

_FALLBACK_HELP_DEBT_STATS = {
    # ATO Taxation Statistics 2025-26 — HELP/HECS debt
    "by_age": {
        "under_25": {"prevalence": 0.35, "mean_balance": 18_500},
        "25_34": {"prevalence": 0.55, "mean_balance": 26_000},
        "35_44": {"prevalence": 0.30, "mean_balance": 22_000},
        "45_54": {"prevalence": 0.12, "mean_balance": 15_000},
        "55_plus": {"prevalence": 0.05, "mean_balance": 10_000},
    },
    "repayment_thresholds": {
        54_435: 0.00,
        62_851: 0.01,
        66_621: 0.02,
        70_619: 0.025,
        74_856: 0.03,
        79_347: 0.035,
        84_108: 0.04,
        89_155: 0.045,
        94_504: 0.05,
        100_175: 0.055,
        106_186: 0.06,
        112_560: 0.065,
        119_320: 0.07,
        126_491: 0.075,
        134_099: 0.08,
        142_173: 0.085,
        150_741: 0.09,
        159_834: 0.095,
        169_486: 0.10,
    },
    "source": "fallback (ATO 2025-26 HELP thresholds)",
}


class RealWorldBenchmarks:
    """Fetches distribution-level lending benchmarks from Australian public sources.

    All methods return dicts with a 'source' key indicating data provenance.
    Every method has a hardcoded fallback so generation never breaks.
    """

    # RBA cash rate target (%) — sourced from RBA Table A2
    RBA_CASH_RATE_HISTORY = {
        (2022, 1): 0.10,  # Pre-tightening
        (2022, 2): 0.85,  # First hikes May+Jun 2022
        (2022, 3): 1.85,  # Jul+Aug+Sep 2022
        (2022, 4): 3.10,  # Oct+Nov+Dec 2022
        (2023, 1): 3.35,  # Feb 2023
        (2023, 2): 3.85,  # Mar+May+Jun 2023
        (2023, 3): 4.10,  # Aug 2023
        (2023, 4): 4.35,  # Nov 2023
        (2024, 1): 4.35,  # Hold
        (2024, 2): 4.35,  # Hold
        (2024, 3): 4.35,  # Hold
        (2024, 4): 4.35,  # Hold
        (2025, 1): 4.10,  # Feb 2025 cut
        (2025, 2): 3.85,  # Projected cuts
        (2025, 3): 3.60,  # Projected
        (2025, 4): 3.35,  # Projected
    }

    # Seasonal lending volume index (1.0 = average month)
    # Source: ABS 5601.0 monthly lending commitments, 5-year average pattern
    SEASONAL_LENDING_INDEX = {
        1: 0.85,  # Jan — summer holidays, low activity
        2: 0.92,  # Feb — market reopens
        3: 1.05,  # Mar — Q1 refinancing push
        4: 1.02,  # Apr — post-Easter
        5: 1.00,  # May — average
        6: 0.95,  # Jun — EOFY, some rush
        7: 0.88,  # Jul — winter lull
        8: 0.93,  # Aug — picking up
        9: 1.05,  # Sep — spring starts
        10: 1.15,  # Oct — peak spring
        11: 1.12,  # Nov — spring continues
        12: 1.08,  # Dec — pre-Christmas rush then drop
    }

    # Cumulative default probability curve by months-on-book
    # Peaks around month 18-24 (seasoning), flattens after month 36
    # Calibrated so terminal rate ≈ 1.04% (APRA Sep Q 2025 NPL)
    DEFAULT_HAZARD_CURVE = {
        3: 0.0010,  # 0.10% — very early defaults (fraud/misrep)
        6: 0.0025,  # 0.25%
        9: 0.0045,  # 0.45%
        12: 0.0065,  # 0.65%
        15: 0.0080,  # 0.80% — approaching seasoning peak
        18: 0.0090,  # 0.90%
        21: 0.0097,  # 0.97%
        24: 0.0100,  # 1.00% — seasoning plateau
        30: 0.0103,  # 1.03%
        36: 0.0104,  # 1.04% — terminal rate
        48: 0.0104,  # Flat after seasoning
        60: 0.0104,
    }

    # Monthly state transition probabilities
    # States: performing, 30dpd, 60dpd, 90dpd, default, prepaid
    # Source: Moody's Australian RMBS performance indices, S&P APAC
    # structured finance roll-rate studies, calibrated to APRA NPL 1.04%.
    #
    # Key calibration points (verified against published data):
    # - CPR 15-22% annually → ~0.015/month prepayment (Moody's AU RMBS)
    # - 30dpd cure rate 40-60% (S&P APAC, Moody's idealised)
    # - 90dpd stays at 90dpd for 6-18 months before resolution (APRA APS 220)
    # - Default = 90+ DPD under APRA; direct 30/60→default is near-zero
    #   for prime RMBS (only bankruptcy/fraud cases skip states)
    ARREARS_TRANSITION_MATRIX = {
        "performing": {"performing": 0.982, "30dpd": 0.003, "prepaid": 0.015},
        "30dpd": {"performing": 0.50, "30dpd": 0.18, "60dpd": 0.31, "default": 0.01},
        "60dpd": {"performing": 0.15, "30dpd": 0.10, "60dpd": 0.15, "90dpd": 0.58, "default": 0.02},
        "90dpd": {"performing": 0.05, "60dpd": 0.05, "90dpd": 0.55, "default": 0.35},
        "default": {"default": 1.0},
        "prepaid": {"prepaid": 1.0},
    }

    def __init__(self):
        self.timeout = httpx.Timeout(20.0, connect=5.0)
        self._cache: dict = {}
        self._cache_timestamps: dict[str, datetime] = {}

    # ------------------------------------------------------------------
    # Caching (same pattern as MacroDataService)
    # ------------------------------------------------------------------

    def _fetch_with_cache(self, cache_key: str, fetch_fn, fallback, ttl_hours: int = _CACHE_TTL_HOURS):
        """Generic cache-or-fetch pattern with TTL and fallback."""
        now = datetime.now(UTC)
        ts = self._cache_timestamps.get(cache_key)
        if ts and cache_key in self._cache:
            if (now - ts) < timedelta(hours=ttl_hours):
                return self._cache[cache_key]

        try:
            value = fetch_fn()
            if value is not None:
                self._cache[cache_key] = value
                self._cache_timestamps[cache_key] = now
                return value
        except Exception as exc:
            logger.warning("Fetch failed for %s: %s — using fallback", cache_key, exc)

        # Serve stale cache if available
        if cache_key in self._cache:
            logger.info("Serving stale cache for %s", cache_key)
            return self._cache[cache_key]

        logger.info("Using hardcoded fallback for %s", cache_key)
        return fallback

    # ------------------------------------------------------------------
    # Public getters
    # ------------------------------------------------------------------

    def get_income_percentiles(self, state: str = "national") -> dict:
        """Fetch income distribution percentiles P10/P25/P50/P75/P90.

        Source: ABS Average Weekly Earnings (cat 6302.0) via SDMX API.
        Falls back to ATO 2022-23 / ABS Aug 2025 values.
        """
        fallback = _FALLBACK_INCOME_PERCENTILES.get(state, _FALLBACK_INCOME_PERCENTILES["national"])
        return self._fetch_with_cache(
            f"income_percentiles_{state}",
            lambda: self._fetch_income_percentiles(state),
            fallback,
        )

    def get_avg_loan_sizes(self) -> dict:
        """Fetch average new loan commitments by purpose.

        Source: ABS Lending Indicators (cat 5601.0) via SDMX API.
        Falls back to ABS Dec Q 2025 values.
        """
        return self._fetch_with_cache(
            "avg_loan_sizes",
            self._fetch_avg_loan_sizes,
            _FALLBACK_AVG_LOAN_SIZES,
        )

    def get_apra_arrears(self) -> dict:
        """Fetch APRA quarterly arrears and distribution data.

        Source: APRA Quarterly ADI Statistics (no REST API — hardcoded
        with audit trail, updated quarterly when new data is published).
        """
        # APRA publishes XLSX only — no public REST API.
        # We maintain a transcribed lookup table (same approach as
        # MacroDataService.APRA_QUARTERLY_BENCHMARKS).
        return self._fetch_with_cache(
            "apra_arrears",
            self._fetch_apra_arrears,
            _FALLBACK_APRA_ARREARS,
        )

    def get_lending_rates(self) -> dict:
        """Fetch current lending rates from RBA Statistical Tables.

        Source: RBA Table F5 — Indicator Lending Rates (CSV download).
        Falls back to approximate 2025/2026 values.
        """
        return self._fetch_with_cache(
            "lending_rates",
            self._fetch_lending_rates,
            _FALLBACK_LENDING_RATES,
        )

    def get_credit_score_distributions(self) -> dict:
        """Return credit score distribution parameters by age bracket.

        Source: Equifax published reports (hardcoded with audit trail,
        updated quarterly when new data is published).
        """
        # Equifax does not publish a public API for score distributions.
        # Values are transcribed from their quarterly published reports.
        return _FALLBACK_CREDIT_SCORE_DISTRIBUTIONS

    def get_approval_rates(self) -> dict:
        """Return estimated approval rates by lender type.

        Derived from ABS lending volumes and APRA ADI statistics.
        """
        return _FALLBACK_APPROVAL_RATES

    def get_sa4_unemployment(self) -> dict:
        """Fetch SA4-level unemployment rates from ABS Labour Force Survey.

        Source: ABS Labour Force (cat 6202.0) via SDMX API — SA4 regions.
        Falls back to hardcoded ~20 key SA4 regions covering all states.
        """
        return self._fetch_with_cache(
            "sa4_unemployment",
            self._fetch_sa4_unemployment,
            _FALLBACK_SA4_UNEMPLOYMENT,
        )

    def get_industry_income_multipliers(self) -> dict:
        """Fetch income multipliers by ANZSIC industry division.

        Source: ABS Average Weekly Earnings (cat 6302.0) via SDMX API.
        Returns multipliers relative to national median income.
        Falls back to ABS AWE 2025 derived values.
        """
        return self._fetch_with_cache(
            "industry_income_multipliers",
            self._fetch_industry_income_multipliers,
            _FALLBACK_INDUSTRY_INCOME_MULTIPLIERS,
        )

    def get_rba_f6_rates(self) -> dict:
        """Fetch housing lending rates from RBA Statistical Table F6.

        Source: RBA Table F6 — Housing Lending Rates (CSV download).
        Falls back to approximate 2025/2026 values.
        """
        return self._fetch_with_cache(
            "f6_rates",
            self._fetch_rba_f6_rates,
            _FALLBACK_F6_RATES,
        )

    def get_help_debt_statistics(self) -> dict:
        """Fetch HELP/HECS debt statistics from ATO Taxation Statistics.

        Source: ATO Taxation Statistics (XLSX download) — HELP debt by
        age bracket and repayment thresholds for 2025-26.
        Falls back to ATO 2025-26 published thresholds.
        """
        return self._fetch_with_cache(
            "help_debt_stats",
            self._fetch_help_debt_statistics,
            _FALLBACK_HELP_DEBT_STATS,
        )

    def get_rba_household_debt(self) -> dict:
        """Fetch household debt ratios from RBA Table E2 (CSV download).

        Source: RBA Statistical Table E2 — Household Finances: Selected Ratios.
        Falls back to RBA Dec Q 2025 values.
        """
        return self._fetch_with_cache(
            "rba_household_debt",
            self._fetch_rba_e2_csv,
            _FALLBACK_RBA_HOUSEHOLD_DEBT,
        )

    def get_calibration_snapshot(self) -> dict:
        """Fetch all benchmarks in one call.

        Returns a complete dict suitable for passing to
        DataGenerator(benchmarks=snapshot).
        """
        snapshot = {
            "income_percentiles": self.get_income_percentiles("national"),
            "income_percentiles_by_state": {
                s: self.get_income_percentiles(s) for s in ["NSW", "VIC", "QLD", "WA", "SA", "TAS", "ACT", "NT"]
            },
            "avg_loan_sizes": self.get_avg_loan_sizes(),
            "apra_arrears": self.get_apra_arrears(),
            "lending_rates": self.get_lending_rates(),
            "credit_score_distributions": self.get_credit_score_distributions(),
            "approval_rates": self.get_approval_rates(),
            "sa4_unemployment": self.get_sa4_unemployment(),
            "industry_income_multipliers": self.get_industry_income_multipliers(),
            "f6_rates": self.get_rba_f6_rates(),
            "help_debt_stats": self.get_help_debt_statistics(),
            "rba_household_debt": self.get_rba_household_debt(),
            "fetched_at": datetime.now(UTC).isoformat(),
        }
        logger.info(
            "Calibration snapshot assembled — sources: %s",
            {k: v.get("source", "n/a") if isinstance(v, dict) else "composite" for k, v in snapshot.items()},
        )
        return snapshot

    # ------------------------------------------------------------------
    # Temporal benchmark lookups
    # ------------------------------------------------------------------

    @classmethod
    def get_cash_rate(cls, year: int, quarter: int) -> float:
        """Return RBA cash rate for a given year/quarter. Falls back to nearest available."""
        key = (year, quarter)
        if key in cls.RBA_CASH_RATE_HISTORY:
            return cls.RBA_CASH_RATE_HISTORY[key]
        # Fallback: nearest quarter
        available = sorted(cls.RBA_CASH_RATE_HISTORY.keys())
        closest = min(available, key=lambda k: abs((k[0] * 4 + k[1]) - (year * 4 + quarter)))
        return cls.RBA_CASH_RATE_HISTORY[closest]

    @classmethod
    def get_seasonal_factor(cls, month: int) -> float:
        """Return seasonal lending volume multiplier for a given month (1-12)."""
        return cls.SEASONAL_LENDING_INDEX.get(month, 1.0)

    @classmethod
    def get_cumulative_default_prob(cls, months_on_book: int) -> float:
        """Return cumulative default probability for a given months-on-book.
        Interpolates linearly between defined points."""
        if months_on_book <= 0:
            return 0.0
        breakpoints = sorted(cls.DEFAULT_HAZARD_CURVE.keys())
        if months_on_book >= breakpoints[-1]:
            return cls.DEFAULT_HAZARD_CURVE[breakpoints[-1]]
        # Linear interpolation
        for i in range(len(breakpoints) - 1):
            if breakpoints[i] <= months_on_book <= breakpoints[i + 1]:
                lo, hi = breakpoints[i], breakpoints[i + 1]
                frac = (months_on_book - lo) / (hi - lo)
                return cls.DEFAULT_HAZARD_CURVE[lo] + frac * (
                    cls.DEFAULT_HAZARD_CURVE[hi] - cls.DEFAULT_HAZARD_CURVE[lo]
                )
        return cls.DEFAULT_HAZARD_CURVE[breakpoints[-1]]

    @classmethod
    def get_transition_probs(cls, current_state: str) -> dict:
        """Return transition probabilities from current loan state."""
        return cls.ARREARS_TRANSITION_MATRIX.get(current_state, {"default": 1.0})

    # ------------------------------------------------------------------
    # Internal fetch methods
    # ------------------------------------------------------------------

    def _fetch_abs_data(self, dataflow_id: str, key: str) -> dict:
        """Fetch from ABS Data API (SDMX-JSON format). No API key required."""
        url = f"https://data.api.abs.gov.au/rest/data/{dataflow_id}/{key}"
        headers = {"Accept": "application/vnd.sdmx.data+json;version=2.0.0"}
        logger.info("Fetching ABS data: %s", url)
        with httpx.Client(timeout=self.timeout) as client:
            response = client.get(url, headers=headers)
            response.raise_for_status()
            return response.json()

    def _parse_abs_latest_value(self, data: dict) -> float | None:
        """Extract the most recent observation from an ABS SDMX-JSON response."""
        try:
            datasets = data.get("data", {}).get("dataSets", [])
            if not datasets:
                return None

            observations = datasets[0].get("observations", {})
            if not observations:
                series = datasets[0].get("series", {})
                if series:
                    first_series = next(iter(series.values()), {})
                    observations = first_series.get("observations", {})

            if not observations:
                return None

            last_key = max(observations.keys(), key=lambda k: int(k.split(":")[-1]))
            value_array = observations[last_key]
            if value_array and value_array[0] is not None:
                return float(value_array[0])
        except (KeyError, IndexError, TypeError, ValueError) as exc:
            logger.warning("Failed to parse ABS SDMX-JSON response: %s", exc)
        return None

    def _parse_abs_series_values(self, data: dict) -> list[float]:
        """Extract all observation values from an ABS SDMX-JSON response."""
        values = []
        try:
            datasets = data.get("data", {}).get("dataSets", [])
            if not datasets:
                return values

            observations = datasets[0].get("observations", {})
            if not observations:
                series = datasets[0].get("series", {})
                if series:
                    first_series = next(iter(series.values()), {})
                    observations = first_series.get("observations", {})

            sorted_keys = sorted(observations.keys(), key=lambda k: int(k.split(":")[-1]))
            for key in sorted_keys:
                val = observations[key]
                if val and val[0] is not None:
                    values.append(float(val[0]))
        except (KeyError, IndexError, TypeError, ValueError) as exc:
            logger.warning("Failed to parse ABS series values: %s", exc)
        return values

    def _fetch_income_percentiles(self, state: str) -> dict | None:
        """Fetch income data from ABS Average Weekly Earnings (cat 6302.0).

        Verified key structure (March 2026):
          Dataflow: AWE (short ID)
          Key: MEASURE.ESTIMATE_TYPE.SEX.SECTOR.INDUSTRY.TSEST.REGION.FREQ
          1 = All employees avg weekly total earnings, 1 = Earnings,
          3 = Persons, 7 = Private+Public, TOT = All Industries,
          10 = Original, region code, S = Half-yearly
        """
        from .macro_data_service import _ABS_STATE_CODES

        state_code = _ABS_STATE_CODES.get(state, "AUS")
        sdmx_key = f"1.1.3.7.TOT.10.{state_code}.S"
        try:
            data = self._fetch_abs_data("AWE", sdmx_key)
            avg_weekly = self._parse_abs_latest_value(data)
            if avg_weekly is None:
                return None

            # Convert weekly to annual
            median_annual = avg_weekly * 52

            # Derive percentiles from median using typical AU income distribution shape
            # (right-skewed lognormal, sigma ≈ 0.65 based on ATO data)
            # These ratios are stable across years (ATO Taxation Statistics)
            return {
                "P10": round(median_annual * 0.337),  # ~$25K at $74K median
                "P25": round(median_annual * 0.567),  # ~$42K
                "P50": round(median_annual),  # median
                "P75": round(median_annual * 1.417),  # ~$105K
                "P90": round(median_annual * 2.091),  # ~$155K
                "source": f"ABS AWE cat 6302.0 (state={state}, derived)",
            }
        except Exception as exc:
            logger.warning("ABS income fetch failed for %s: %s", state, exc)
            return None

    def _fetch_avg_loan_sizes(self) -> dict | None:
        """Fetch average loan sizes from ABS Lending Indicators (cat 5601.0).

        ABS Lending Indicators dataflow: ABS,LENDING_INDICATORS_M
        Key structure varies — we fetch the headline new loan commitments.
        """
        try:
            # Owner-occupier new loan commitments — average size
            # Measure 22 = average loan size, housing finance
            data = self._fetch_abs_data("ABS,LENDING_INDICATORS_M", "22.10.Q")
            owner_occ = self._parse_abs_latest_value(data)

            if owner_occ is None:
                return None

            # Derive other segments from owner-occupier using stable ratios
            # (ABS Dec Q 2025 ratios)
            return {
                "owner_occupier": round(owner_occ),
                "first_home_buyer": round(owner_occ * 0.808),  # $560K / $693K
                "investor": round(owner_occ * 0.988),  # $685K / $693K
                "personal": _FALLBACK_AVG_LOAN_SIZES["personal"],  # ABS personal lending
                "business": _FALLBACK_AVG_LOAN_SIZES["business"],  # ABS business lending
                "source": f"ABS Lending Indicators cat 5601.0 (owner_occ={owner_occ:.0f}, derived)",
            }
        except Exception as exc:
            logger.warning("ABS lending indicators fetch failed: %s", exc)
            return None

    def _fetch_apra_arrears(self) -> dict | None:
        """Fetch latest APRA quarterly arrears data.

        APRA publishes via XLSX (no REST API). We try to download the
        latest quarterly statistics page, but fall back to hardcoded
        values if the download or parse fails.
        """
        try:
            # Try to fetch the APRA quarterly statistics summary page
            url = "https://www.apra.gov.au/quarterly-authorised-deposit-taking-institution-statistics"
            with httpx.Client(timeout=self.timeout, follow_redirects=True) as client:
                response = client.get(url)
                response.raise_for_status()
                html = response.text

            # Look for the latest XLSX download link
            # APRA page structure: links to .xlsx files with quarterly data
            import re

            xlsx_links = re.findall(
                r'href="(https?://[^"]*(?:quarterly|qadi)[^"]*\.xlsx)"',
                html,
                re.IGNORECASE,
            )

            if not xlsx_links:
                logger.info("No APRA XLSX links found on page — using fallback")
                return None

            # Download the most recent XLSX (capped at 50MB to prevent DoS)
            xlsx_url = xlsx_links[0]
            # Validate URL domain before fetching
            from urllib.parse import urlparse

            parsed = urlparse(xlsx_url)
            if parsed.hostname and not parsed.hostname.endswith(".gov.au"):
                logger.warning("APRA XLSX URL has non-gov.au domain: %s", parsed.hostname)
                return None
            logger.info("Downloading APRA XLSX: %s", xlsx_url)
            _MAX_XLSX_SIZE = 50 * 1024 * 1024  # 50 MB
            with httpx.Client(timeout=httpx.Timeout(30.0, connect=10.0), follow_redirects=True) as client:
                xlsx_response = client.get(xlsx_url)
                xlsx_response.raise_for_status()
                if len(xlsx_response.content) > _MAX_XLSX_SIZE:
                    logger.warning("APRA XLSX too large (%d bytes) — skipped", len(xlsx_response.content))
                    return None

            return self._parse_apra_xlsx(xlsx_response.content)

        except Exception as exc:
            logger.warning("APRA fetch failed: %s — using fallback", exc)
            return None

    def _parse_apra_xlsx(self, xlsx_bytes: bytes) -> dict | None:
        """Parse APRA quarterly ADI statistics XLSX for arrears data."""
        try:
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)

            # APRA XLSX typically has sheets like "Table 1", "Table 2", etc.
            # Property exposures are usually in "Table 1b" or similar
            result = dict(_FALLBACK_APRA_ARREARS)  # start with fallback structure

            matched_sheet = None
            for sheet_name in wb.sheetnames:
                sheet_lower = sheet_name.lower()
                if "property" in sheet_lower or "housing" in sheet_lower:
                    matched_sheet = sheet_name
                    ws = wb[sheet_name]
                    # Scan for NPL rate, arrears, LVR distribution
                    for row in ws.iter_rows(min_row=1, max_row=100, values_only=True):
                        if not row or not row[0]:
                            continue
                        label = str(row[0]).lower().strip()

                        if "non-performing" in label and "rate" in label:
                            for cell in row[1:]:
                                if isinstance(cell, (int, float)) and 0 < cell < 1:
                                    result["npl_rate"] = round(cell, 4)
                                    break

                        if "90" in label and "arrears" in label:
                            for cell in row[1:]:
                                if isinstance(cell, (int, float)) and 0 < cell < 1:
                                    result["total_arrears_rate"] = round(cell, 4)
                                    break

            result["source"] = f"APRA XLSX ({matched_sheet or 'unknown sheet'})"
            return result

        except Exception as exc:
            logger.warning("APRA XLSX parse failed: %s", type(exc).__name__)
            return None

        finally:
            try:
                wb.close()
            except Exception:
                pass

    def _fetch_lending_rates(self) -> dict | None:
        """Fetch lending rates from RBA Table F5 (CSV download).

        RBA publishes CSV files at: https://www.rba.gov.au/statistics/tables/
        Table F5: Indicator Lending Rates
        """
        try:
            url = "https://www.rba.gov.au/statistics/tables/csv/f5-data.csv"
            with httpx.Client(timeout=self.timeout, follow_redirects=True) as client:
                response = client.get(url)
                response.raise_for_status()

            return self._parse_rba_f5_csv(response.text)
        except Exception as exc:
            logger.warning("RBA F5 CSV fetch failed: %s", exc)
            return None

    def _parse_rba_f5_csv(self, csv_text: str) -> dict | None:
        """Parse RBA Table F5 CSV for lending rate data."""
        try:
            # RBA CSVs have header rows before the data starts
            lines = csv_text.strip().split("\n")

            # Find the header row (contains 'Series ID' or 'Title')
            header_idx = None
            for i, line in enumerate(lines):
                if "series id" in line.lower() or "title" in line.lower():
                    header_idx = i
                    break

            if header_idx is None:
                return None

            # Parse as CSV from the header row
            reader = csv.reader(lines[header_idx:])
            next(reader)

            # Find columns for variable rate, fixed rates
            # RBA F5 columns include housing rates
            last_data_row = None
            for row in reader:
                if row and row[0] and len(row) > 1:
                    # Skip non-numeric rows
                    try:
                        float(row[1])
                        last_data_row = row
                    except (ValueError, IndexError):
                        continue

            if last_data_row is None:
                return None

            # Extract rates — RBA F5 has standard variable rate in early columns
            # The exact column indices vary; we look for the first few numeric values
            rates = []
            for cell in last_data_row[1:]:
                try:
                    val = float(cell)
                    if 2.0 < val < 15.0:  # plausible lending rate range
                        rates.append(val)
                except (ValueError, TypeError):
                    continue

            if len(rates) >= 2:
                return {
                    "avg_variable_rate": rates[0],
                    "avg_fixed_1yr": rates[1] if len(rates) > 1 else rates[0] - 0.30,
                    "avg_fixed_3yr": rates[2] if len(rates) > 2 else rates[0] - 0.60,
                    "discounted_rate": rates[0] - 0.40,
                    "source": "RBA Table F5 (live)",
                }

        except Exception as exc:
            logger.warning("RBA F5 CSV parse failed: %s", exc)
        return None

    def _fetch_sa4_unemployment(self) -> dict | None:
        """Fetch SA4-level unemployment rates from ABS Labour Force (cat 6202.0).

        ABS Labour Force dataflow: ABS,LF
        Key structure: MEASURE.REGION.SEX.AGE.ADJUSTMENT.FREQ
        We fetch unemployment rate (UNR_RATE) by SA4 region.
        """
        try:
            # Fetch unemployment rate by SA4 region
            # 14 = Unemployment rate, SA4 regions, 3 = Persons, 999 = All ages,
            # 20 = Seasonally adjusted, M = Monthly
            data = self._fetch_abs_data("ABS,LF", "14.*.3.999.20.M")
            datasets = data.get("data", {}).get("dataSets", [])
            if not datasets:
                return None

            series = datasets[0].get("series", {})
            if not series:
                return None

            # Extract SA4 codes and their latest unemployment rates
            # The dimension structure maps series keys to SA4 region codes
            structures = data.get("data", {}).get("structure", {}).get("dimensions", {})
            series_dims = structures.get("series", []) if structures else []

            # Find the region dimension (usually index 1)
            region_dim = None
            for dim in series_dims:
                if "region" in dim.get("id", "").lower() or "sa4" in dim.get("name", "").lower():
                    region_dim = dim
                    break

            if not region_dim:
                logger.warning("Could not find SA4 region dimension in ABS LF response")
                return None

            region_values = region_dim.get("values", [])
            result = {}

            for series_key, series_data in series.items():
                observations = series_data.get("observations", {})
                if not observations:
                    continue

                # Get latest observation
                last_key = max(observations.keys(), key=lambda k: int(k.split(":")[-1]))
                value_array = observations[last_key]
                if not value_array or value_array[0] is None:
                    continue

                # Extract region index from the series key
                key_parts = series_key.split(":")
                if len(key_parts) < 2:
                    continue
                region_idx = int(key_parts[1])

                if region_idx < len(region_values):
                    sa4_code = region_values[region_idx].get("id", "")
                    # Convert percentage to proportion
                    rate = float(value_array[0]) / 100.0
                    if 0 < rate < 0.30:  # plausible unemployment rate range
                        result[sa4_code] = round(rate, 4)

            if result:
                result["source"] = f"ABS Labour Force cat 6202.0 ({len(result)} SA4 regions)"
                logger.info("Fetched SA4 unemployment for %d regions", len(result))
                return result

        except Exception as exc:
            logger.warning("ABS SA4 unemployment fetch failed: %s", exc)
        return None

    def _fetch_industry_income_multipliers(self) -> dict | None:
        """Fetch Average Weekly Earnings by ANZSIC division from ABS (cat 6302.0).

        ABS AWE dataflow: AWE
        Key structure: MEASURE.ESTIMATE_TYPE.SEX.SECTOR.INDUSTRY.TSEST.REGION.FREQ
        We fetch earnings by industry division and compute multipliers relative
        to the national median.
        """
        try:
            # Fetch all-employees avg weekly total earnings by industry
            # 1 = Avg weekly total earnings, 1 = Earnings, 3 = Persons,
            # 7 = Private+Public, * = All industries, 10 = Original,
            # AUS = National, S = Half-yearly
            data = self._fetch_abs_data("AWE", "1.1.3.7.*.10.AUS.S")
            datasets = data.get("data", {}).get("dataSets", [])
            if not datasets:
                return None

            series = datasets[0].get("series", {})
            if not series:
                return None

            # Extract ANZSIC dimension values
            structures = data.get("data", {}).get("structure", {}).get("dimensions", {})
            series_dims = structures.get("series", []) if structures else []

            # Find the industry dimension (usually index 4)
            industry_dim = None
            for dim in series_dims:
                if "industry" in dim.get("id", "").lower():
                    industry_dim = dim
                    break

            if not industry_dim:
                logger.warning("Could not find industry dimension in ABS AWE response")
                return None

            industry_values = industry_dim.get("values", [])
            earnings_by_industry = {}

            for series_key, series_data in series.items():
                observations = series_data.get("observations", {})
                if not observations:
                    continue

                # Get latest observation
                last_key = max(observations.keys(), key=lambda k: int(k.split(":")[-1]))
                value_array = observations[last_key]
                if not value_array or value_array[0] is None:
                    continue

                # Extract industry index from the series key
                key_parts = series_key.split(":")
                if len(key_parts) < 5:
                    continue
                industry_idx = int(key_parts[4])

                if industry_idx < len(industry_values):
                    industry_id = industry_values[industry_idx].get("id", "")
                    earnings_by_industry[industry_id] = float(value_array[0])

            if not earnings_by_industry:
                return None

            # Find the "TOT" (all industries) value as the national median baseline
            national_earnings = earnings_by_industry.pop("TOT", None)
            if national_earnings is None:
                # Use mean of all industries as baseline
                national_earnings = sum(earnings_by_industry.values()) / len(earnings_by_industry)

            # Compute multipliers relative to national
            result = {}
            # Map ABS industry codes to ANZSIC division letters
            anzsic_map = {
                "A": "A",
                "B": "B",
                "C": "C",
                "D": "D",
                "E": "E",
                "F": "F",
                "G": "G",
                "H": "H",
                "I": "I",
                "J": "J",
                "K": "K",
                "L": "L",
                "M": "M",
                "N": "N",
                "O": "O",
                "P": "P",
                "Q": "Q",
                "R": "R",
                "S": "S",
            }

            for code, earnings in earnings_by_industry.items():
                division = anzsic_map.get(code, code)
                multiplier = round(earnings / national_earnings, 2)
                if 0.2 < multiplier < 5.0:  # plausible range
                    result[division] = multiplier

            if result:
                result["source"] = f"ABS AWE cat 6302.0 ({len(result)} ANZSIC divisions)"
                logger.info("Fetched industry income multipliers for %d divisions", len(result))
                return result

        except Exception as exc:
            logger.warning("ABS industry income multipliers fetch failed: %s", exc)
        return None

    def _fetch_rba_f6_rates(self) -> dict | None:
        """Fetch housing lending rates from RBA Table F6 (CSV download).

        RBA publishes CSV files at: https://www.rba.gov.au/statistics/tables/
        Table F6: Housing Lending Rates
        """
        try:
            url = "https://www.rba.gov.au/statistics/tables/csv/f6-data.csv"
            with httpx.Client(timeout=self.timeout, follow_redirects=True) as client:
                response = client.get(url)
                response.raise_for_status()

            return self._parse_rba_f6_csv(response.text)
        except Exception as exc:
            logger.warning("RBA F6 CSV fetch failed: %s", exc)
            return None

    def _parse_rba_f6_csv(self, csv_text: str) -> dict | None:
        """Parse RBA Table F6 CSV for housing lending rate data."""
        try:
            # RBA CSVs have header rows before the data starts
            lines = csv_text.strip().split("\n")

            # Find the header row (contains 'Series ID' or 'Title')
            header_idx = None
            for i, line in enumerate(lines):
                if "series id" in line.lower() or "title" in line.lower():
                    header_idx = i
                    break

            if header_idx is None:
                return None

            # Parse as CSV from the header row
            reader = csv.reader(lines[header_idx:])
            headers = next(reader)

            # Find columns for owner-occupier and investor rates
            oo_var_col = None
            oo_fixed_col = None
            inv_var_col = None
            inv_fixed_col = None

            for idx, header in enumerate(headers):
                header_lower = header.lower()
                if "owner" in header_lower and "variable" in header_lower:
                    oo_var_col = idx
                elif "owner" in header_lower and "fixed" in header_lower and "3" in header_lower:
                    oo_fixed_col = idx
                elif "investor" in header_lower and "variable" in header_lower:
                    inv_var_col = idx
                elif "investor" in header_lower and "fixed" in header_lower and "3" in header_lower:
                    inv_fixed_col = idx

            # Get the last data row
            last_data_row = None
            for row in reader:
                if row and row[0] and len(row) > 1:
                    try:
                        float(row[1])
                        last_data_row = row
                    except (ValueError, IndexError):
                        continue

            if last_data_row is None:
                return None

            # Extract rates — use column indices if found, otherwise positional
            def _safe_float(row, col_idx):
                if col_idx is not None and col_idx < len(row):
                    try:
                        val = float(row[col_idx])
                        if 2.0 < val < 15.0:  # plausible lending rate range
                            return val
                    except (ValueError, TypeError):
                        pass
                return None

            # Collect all plausible rates as positional fallback
            rates = []
            for cell in last_data_row[1:]:
                try:
                    val = float(cell)
                    if 2.0 < val < 15.0:
                        rates.append(val)
                except (ValueError, TypeError):
                    continue

            oo_var = _safe_float(last_data_row, oo_var_col)
            oo_fixed = _safe_float(last_data_row, oo_fixed_col)
            inv_var = _safe_float(last_data_row, inv_var_col)
            inv_fixed = _safe_float(last_data_row, inv_fixed_col)

            # Fall back to positional if column matching failed
            if oo_var is None and len(rates) >= 1:
                oo_var = rates[0]
            if oo_fixed is None and len(rates) >= 2:
                oo_fixed = rates[1]
            if inv_var is None and len(rates) >= 3:
                inv_var = rates[2]
            if inv_fixed is None and len(rates) >= 4:
                inv_fixed = rates[3]

            if oo_var is not None:
                return {
                    "owner_occupier_variable": oo_var,
                    "owner_occupier_fixed_3yr": oo_fixed or round(oo_var - 0.38, 2),
                    "investor_variable": inv_var or round(oo_var + 0.31, 2),
                    "investor_fixed_3yr": inv_fixed or round(oo_var - 0.12, 2),
                    "source": "RBA Table F6 (live)",
                }

        except Exception as exc:
            logger.warning("RBA F6 CSV parse failed: %s", exc)
        return None

    def _fetch_rba_e2_csv(self) -> dict | None:
        """Fetch household debt ratios from RBA Table E2 (CSV download).

        RBA publishes CSV files at: https://www.rba.gov.au/statistics/tables/
        Table E2: Household Finances — Selected Ratios
        """
        try:
            url = "https://www.rba.gov.au/statistics/tables/csv/e2-data.csv"
            with httpx.Client(timeout=self.timeout, follow_redirects=True) as client:
                response = client.get(url)
                response.raise_for_status()

            return self._parse_rba_e2_csv(response.text)
        except Exception as exc:
            logger.warning("RBA E2 CSV fetch failed: %s", exc)
            return None

    def _parse_rba_e2_csv(self, csv_text: str) -> dict | None:
        """Parse RBA Table E2 CSV for household debt ratio data."""
        try:
            lines = csv_text.strip().split("\n")

            # Find header row (contains 'Series ID' or 'Title')
            header_idx = None
            title_idx = None
            for i, line in enumerate(lines):
                lower = line.lower()
                if "series id" in lower:
                    header_idx = i
                elif "title" in lower and title_idx is None:
                    title_idx = i

            if header_idx is None and title_idx is None:
                return None

            # Initialise column indices before conditional block to avoid
            # UnboundLocalError when title_idx is None but header_idx is set
            housing_dti_col = None
            total_dti_col = None
            dta_col = None
            interest_col = None

            # Use title row to identify columns by name
            if title_idx is not None:
                title_reader = csv.reader([lines[title_idx]])
                titles = next(title_reader)
                titles_lower = [t.lower().strip() for t in titles]

                for j, title in enumerate(titles_lower):
                    if "housing" in title and "debt" in title and "income" in title:
                        housing_dti_col = j
                    elif "total" in title and "debt" in title and "income" in title:
                        total_dti_col = j
                    elif "debt" in title and "assets" in title:
                        dta_col = j
                    elif "interest" in title and "income" in title:
                        interest_col = j

            # Find last data row
            start = (header_idx or title_idx or 0) + 1
            reader = csv.reader(lines[start:])
            last_data_row = None
            for row in reader:
                if row and row[0] and len(row) > 1:
                    try:
                        float(row[1])
                        last_data_row = row
                    except (ValueError, IndexError):
                        continue

            if last_data_row is None:
                return None

            def _safe(row, col):
                if col is not None and col < len(row):
                    try:
                        return float(row[col])
                    except (ValueError, TypeError):
                        pass
                return None

            housing_dti = _safe(last_data_row, housing_dti_col)
            total_dti = _safe(last_data_row, total_dti_col)
            dta = _safe(last_data_row, dta_col)
            interest_ratio = _safe(last_data_row, interest_col)

            # Need at least one ratio to be useful
            if housing_dti is None and total_dti is None:
                return None

            # RBA E2 uses percentage format (e.g. 141.0 = 141%).
            # Normalize to ratio format (1.41) to match our fallback values.
            def _pct_to_ratio(val, fallback):
                if val is None:
                    return fallback
                return val / 100.0 if val > 5.0 else val

            return {
                "housing_debt_to_income": _pct_to_ratio(housing_dti, 1.41),
                "total_debt_to_income": _pct_to_ratio(total_dti, 1.87),
                "debt_to_assets": _pct_to_ratio(dta, 0.20),
                "interest_payments_to_income": _pct_to_ratio(interest_ratio, 0.098),
                "quarter": last_data_row[0] if last_data_row else "unknown",
                "source": "RBA Table E2 (live)",
            }

        except Exception as exc:
            logger.warning("RBA E2 CSV parse failed: %s", exc)
        return None

    def _fetch_help_debt_statistics(self) -> dict | None:
        """Fetch HELP/HECS debt statistics from ATO Taxation Statistics.

        ATO publishes XLSX files with HELP debt data. We try to download
        the latest statistics, but fall back to hardcoded values if the
        download or parse fails.
        """
        try:
            # Try to fetch the ATO Taxation Statistics page for HELP data
            url = "https://www.ato.gov.au/about-ato/research-and-statistics/in-detail/taxation-statistics"
            with httpx.Client(timeout=self.timeout, follow_redirects=True) as client:
                response = client.get(url)
                response.raise_for_status()
                html = response.text

            # Look for XLSX download links related to HELP debt
            import re

            xlsx_links = re.findall(
                r'href="(https?://[^"]*(?:help|hecs|higher-education)[^"]*\.xlsx)"',
                html,
                re.IGNORECASE,
            )

            if not xlsx_links:
                logger.info("No ATO HELP XLSX links found on page — using fallback")
                return None

            # Download the most recent XLSX (capped at 50MB to prevent DoS)
            xlsx_url = xlsx_links[0]
            from urllib.parse import urlparse

            parsed = urlparse(xlsx_url)
            if parsed.hostname and not parsed.hostname.endswith(".gov.au"):
                logger.warning("ATO XLSX URL has non-gov.au domain: %s", parsed.hostname)
                return None
            logger.info("Downloading ATO HELP XLSX: %s", xlsx_url)
            _MAX_XLSX_SIZE = 50 * 1024 * 1024
            with httpx.Client(timeout=httpx.Timeout(30.0, connect=10.0), follow_redirects=True) as client:
                xlsx_response = client.get(xlsx_url)
                xlsx_response.raise_for_status()
                if len(xlsx_response.content) > _MAX_XLSX_SIZE:
                    logger.warning("ATO XLSX too large (%d bytes) — skipped", len(xlsx_response.content))
                    return None

            return self._parse_help_debt_xlsx(xlsx_response.content)

        except Exception as exc:
            logger.warning("ATO HELP debt fetch failed: %s — using fallback", exc)
            return None

    def _parse_help_debt_xlsx(self, xlsx_bytes: bytes) -> dict | None:
        """Parse ATO HELP debt statistics XLSX for debt by age bracket."""
        try:
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)

            # Start with fallback structure, overlay any parsed values
            result = dict(_FALLBACK_HELP_DEBT_STATS)
            result["by_age"] = dict(_FALLBACK_HELP_DEBT_STATS["by_age"])

            # ATO XLSX typically has sheets with HELP debt data by age
            matched_help_sheet = None
            for sheet_name in wb.sheetnames:
                sheet_lower = sheet_name.lower()
                if "help" in sheet_lower or "hecs" in sheet_lower or "debt" in sheet_lower:
                    matched_help_sheet = sheet_name
                    ws = wb[sheet_name]
                    for row in ws.iter_rows(min_row=1, max_row=200, values_only=True):
                        if not row or not row[0]:
                            continue
                        label = str(row[0]).lower().strip()

                        # Try to match age brackets
                        age_mapping = {
                            "under 25": "under_25",
                            "25-34": "25_34",
                            "25 to 34": "25_34",
                            "35-44": "35_44",
                            "35 to 44": "35_44",
                            "45-54": "45_54",
                            "45 to 54": "45_54",
                            "55+": "55_plus",
                            "55 and over": "55_plus",
                        }

                        for pattern, age_key in age_mapping.items():
                            if pattern in label:
                                # Look for prevalence and mean balance values
                                numeric_vals = []
                                for cell in row[1:]:
                                    if isinstance(cell, (int, float)) and cell > 0:
                                        numeric_vals.append(cell)
                                if len(numeric_vals) >= 2:
                                    prevalence = numeric_vals[0]
                                    mean_bal = numeric_vals[1]
                                    # Validate ranges (security: bound mean_balance)
                                    if 0 < prevalence <= 1.0 and 0 < mean_bal < 200_000:
                                        result["by_age"][age_key] = {
                                            "prevalence": round(prevalence, 2),
                                            "mean_balance": round(mean_bal),
                                        }
                                break

            result["source"] = f"ATO HELP Statistics ({matched_help_sheet or 'unknown sheet'})"
            return result

        except Exception as exc:
            logger.warning("ATO HELP XLSX parse failed: %s", type(exc).__name__)
            return None

        finally:
            try:
                wb.close()
            except Exception:
                pass
